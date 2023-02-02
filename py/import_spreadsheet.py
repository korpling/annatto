from glob import iglob
from graphannis.graph import GraphUpdate
from graphupdate_util import *
import logging
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    import oxidized_importer
    _openpyxl_zip_path = 'py/zip/openpyxl.zip'
    _finder = oxidized_importer.OxidizedZipFinder.from_path(_openpyxl_zip_path)
    sys.meta_path.insert(0, _finder)
    from openpyxl import load_workbook

# constants
_FILE_ENDINGS = ('.xlsx', '.xls', '.ods')
_PROP_COLUMN_MAP = 'column_map'

# logger
_logger = logging.getLogger(__name__)
_stream = logging.StreamHandler(stream=sys.stdout)
_stream.setLevel(logging.INFO)
_logger.setLevel(logging.INFO)
_logger.addHandler(_stream)

def map_spreadsheet(u, path, doc_path, column_map):
    wb = load_workbook(path)
    sh = wb[wb.sheetnames[0]]
    name_to_index = {cell.value: cell.col_idx for cell in next(sh.rows)}
    merged_cells = sh.merged_cells
    row_indices = {col: set(range(2, sh.max_row + 2))   # start from 2, because 1 contains header (1-based indices)
                   for col in range(sh.min_column, sh.max_column + 2)}
    for merged_range in merged_cells.ranges:
        row_indices[merged_range.min_col] -= set(range(merged_range.min_row + 1, merged_range.max_row + 1))  # only keep the first row index (therefore min_row + 1), bc it contains the merged cell value
    rows = list(sh.rows)
    is_multi_tok = len(column_map) > 1
    empty_toks = []
    if is_multi_tok:
        # create empty tokens
        n_empty_toks = sh.max_row + 2
        for i in range(n_empty_toks):
            empty_toks.append((i + 1, map_token(u, doc_path, i + 1, '', ' ', add_annis_layer=False)))
        add_order_relations(u, [id_ for _, id_ in empty_toks], '')
    tok_count = len(empty_toks)
    span_count = 0
    existing_spans = {}
    for tok_name, anno_names in column_map.items():
        tok_map = {}
        col = name_to_index[tok_name]
        tok_row_indices = sorted(row_indices[col]) + [max(row_indices[col]) + 1]
        with open('xlsx_debug.txt', 'a') as f:
            f.write(f'tok row indices for {tok_name}: {tok_row_indices}\n')
        for end_i in range(1, len(tok_row_indices)):
            start = tok_row_indices[end_i - 1]
            end = tok_row_indices[end_i]  # exclusive
            if start - 1 >= len(rows):
                break
            cell = rows[start - 1][col - 1]  # indices are 1-based and must thus be corrected; value is always contained in first cell of merged cell            
            if cell.value is not None and (value := str(cell.value).strip()):
                if is_multi_tok:
                    span_count += 1
                    tok_id = map_token_as_span(u, doc_path, span_count, tok_name, value, start_time=None, end_time=None, empty_toks=empty_toks[start:end])
                    for t in range(start, end):
                        tok_map[tok_id] = t
                    existing_spans[tuple(range(start, end)) + (tok_name,)] = tok_id
                else:
                    tok_count += 1
                    tok_id = map_token(u, doc_path, tok_count, tok_name, value)
                    for t in range(start, end):
                        tok_map[tok_id] = t
        add_order_relations(u, sorted(tok_map, key=lambda k: tok_map[k]), tok_name)
        for anno_name in anno_names:
            split_qname = anno_name.split('::', 1)
            ns, name = split_qname if len(split_qname) == 2 else (tok_name, anno_name)
            col = name_to_index[anno_name]
            available_indices = sorted(row_indices[col]) + [max(row_indices[col]) + 1]
            for end_i in range(1, len(available_indices)):
                start = available_indices[end_i - 1]
                if start - 1 >= len(rows):
                    break
                end = available_indices[end_i]  # exclusive
                covered_toks = sorted({tok_id for tok_id, t in tok_map.items() if start <= t < end}, key=lambda e: tok_map[e])                
                cell = rows[start - 1][col - 1]
                if cell.value is not None and (value := str(cell.value).strip()):
                    span_key = tuple(range(start, end)) + (tok_name,)
                    if span_key not in existing_spans:
                        span_count += 1
                        existing_spans[span_key] = map_annotation(u, doc_path, span_count, ns, name, value, *covered_toks)
                    else:
                        node_id = existing_spans[span_key]
                        map_annotation_to_existing_node(u, node_id, ns, name, value)


def parse_config_s(config_s):
    config = {}
    for group in config_s.strip().split(';'):
        tok_name, anno_spec = group.strip().split('=')
        config[tok_name.strip()] = [a.strip() for a in anno_spec.replace('{', '').replace('}', '').split(',')]
    return config


def start_import(path, **properties):
    u = GraphUpdate()
    if _PROP_COLUMN_MAP not in properties:
        _logger.error(f'You need to configure setting {_PROP_COLUMN_MAP} for importing spreadsheets.')
        raise ValueError
    column_map = parse_config_s(properties[_PROP_COLUMN_MAP])
    for path, internal_path in path_structure(u, path, file_endings=_FILE_ENDINGS):
        map_spreadsheet(u, path, internal_path, column_map)
    return u
